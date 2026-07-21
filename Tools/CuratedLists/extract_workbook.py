#!/usr/bin/env python3
"""One-shot script that reads the Great Books audit workbook and emits:

    gbww-works.json   — 183 normalized GBWW work rows with stable workIDs
    verified-seed.json — 29 hand-verified LibriVox→IA links from the
                         Recording Versions sheet

Run:  python3 Tools/CuratedLists/extract_workbook.py
"""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import openpyxl

WORKBOOK = Path("great_books_librivox_batch_10_completed_through_row13.xlsx")
OUT_WORKS = Path("Tools/CuratedLists/gbww-works.json")
OUT_SEED = Path("Tools/CuratedLists/verified-seed.json")

# ── GBWW second-edition constituents ──────────────────────────────────
# For corpus / grouped entries where the workbook does not enumerate
# every constituent, this mapping supplies the canonical list from the
# second-edition table of contents (1990).  Recorded as of 2026-07-21
# against https://www.bopsecrets.org/gateway/book-lists/greatbooks.htm
# and the Syntopicon volume listings.
#
# Key: "<author>|<title>"  (normalized: lowercased, trimmed)
# Value: list of constituent titles.

GBWW_CONSTITUENTS: dict[str, list[str]] = {
    # ── Vol 4: Aeschylus — Plays ──
    "aeschylus|plays": [
        "Agamemnon",
        "Choephoroe (The Libation Bearers)",
        "Eumenides (The Furies)",
        "The Suppliant Maidens",
        "Prometheus Bound",
        "The Seven Against Thebes",
        "The Persians",
    ],
    # ── Vol 4: Sophocles — Plays ──
    "sophocles|plays": [
        "Oedipus Rex",
        "Oedipus at Colonus",
        "Antigone",
        "Electra",
        "Ajax",
        "Philoctetes",
        "Trachiniae",
    ],
    # ── Vol 4: Euripides — Plays ──
    "euripides|plays": [
        "Rhesus",
        "Medea",
        "Hippolytus",
        "Alcestis",
        "Heracleidae",
        "The Suppliants",
        "Trojan Women",
        "Ion",
        "Helen",
        "Andromache",
        "Electra",
        "Bacchantes",
        "Hecuba",
        "Heracles Mad",
        "Phoenician Women",
        "Orestes",
        "Iphigenia in Tauris",
        "Iphigenia in Aulis",
        "The Cyclops",
    ],
    # ── Vol 4: Aristophanes — Plays ──
    "aristophanes|plays": [
        "The Acharnians",
        "The Knights",
        "The Clouds",
        "The Wasps",
        "Peace",
        "The Birds",
        "The Frogs",
        "Lysistrata",
        "Thesmophoriazusae",
        "Ecclesiazusae",
        "Plutus",
    ],
    # ── Vol 6: Plato — Dialogues ──
    "plato|dialogues": [
        "Charmides",
        "Lysis",
        "Laches",
        "Protagoras",
        "Euthydemus",
        "Cratylus",
        "Phaedrus",
        "Ion",
        "Symposium",
        "Meno",
        "Euthyphro",
        "Apology",
        "Crito",
        "Phaedo",
        "Gorgias",
        "The Republic",
        "Timaeus",
        "Critias",
        "Parmenides",
        "Theaetetus",
        "Sophist",
        "Statesman",
        "Philebus",
        "Laws",
        "The Seventh Letter",
    ],
    # ── Vol 7: Aristotle — Works ──
    "aristotle|works": [
        "Categories",
        "On Interpretation",
        "Prior Analytics",
        "Posterior Analytics",
        "Topics",
        "Sophistical Refutations",
        "Physics",
        "On the Heavens",
        "On Generation and Corruption",
        "Meteorology",
        "Metaphysics",
        "On the Soul",
        "Parva Naturalia",
        "History of Animals",
        "Parts of Animals",
        "Motion of Animals",
        "Gait of Animals",
        "Generation of Animals",
        "Nicomachean Ethics",
        "Politics",
        "The Athenian Constitution",
        "Rhetoric",
        "Poetics",
    ],
    # ── Vol 9: Hippocrates — Works ──
    "hippocrates|works": [
        "The Oath",
        "On Ancient Medicine",
        "On Airs, Waters, and Places",
        "The Book of Prognostics",
        "On Regimen in Acute Diseases",
        "Of the Epidemics",
        "On Injuries of the Head",
        "On the Surgery",
        "On Fractures",
        "On the Articulations",
        "Instruments of Reduction",
        "Aphorisms",
        "The Law",
        "On Ulcers",
        "On Fistulae",
        "On Hemorrhoids",
        "On the Sacred Disease",
    ],
    # ── Vol 10: Archimedes — Works ──
    "archimedes|works": [
        "On the Sphere and Cylinder",
        "Measurement of a Circle",
        "On Conoids and Spheroids",
        "On Spirals",
        "On the Equilibrium of Planes",
        "The Sand-Reckoner",
        "Quadrature of the Parabola",
        "On Floating Bodies",
        "Book of Lemmas",
        "The Method Treating of Mechanical Problems",
    ],
    # ── Vol 17: Plotinus — The Six Enneads ──
    "plotinus|the six enneads": [
        "First Ennead",
        "Second Ennead",
        "Third Ennead",
        "Fourth Ennead",
        "Fifth Ennead",
        "Sixth Ennead",
    ],
    # ── Vol 24,25: Shakespeare — Plays ──
    "william shakespeare|plays": [
        "The First Part of King Henry the Sixth",
        "The Second Part of King Henry the Sixth",
        "The Third Part of King Henry the Sixth",
        "The Tragedy of King Richard the Third",
        "The Comedy of Errors",
        "Titus Andronicus",
        "The Taming of the Shrew",
        "The Two Gentlemen of Verona",
        "Love's Labour's Lost",
        "Romeo and Juliet",
        "The Tragedy of King Richard II",
        "A Midsummer Night's Dream",
        "The Life and Death of King John",
        "The Merchant of Venice",
        "The First Part of King Henry the Fourth",
        "The Second Part of King Henry the Fourth",
        "Much Ado About Nothing",
        "The Life of King Henry the Fifth",
        "Julius Caesar",
        "As You Like It",
        "Twelfth Night; or, What You Will",
        "The Tragedy of Hamlet, Prince of Denmark",
        "The Merry Wives of Windsor",
        "Troilus and Cressida",
        "All's Well That Ends Well",
        "Measure for Measure",
        "Othello, the Moor of Venice",
        "King Lear",
        "Macbeth",
        "Antony and Cleopatra",
        "Coriolanus",
        "Timon of Athens",
        "Pericles, Prince of Tyre",
        "Cymbeline",
        "The Winter's Tale",
        "The Tempest",
        "The Famous History of the Life of King Henry VIII",
    ],
    # ── Vol 26: Shakespeare — Sonnets ──
    "william shakespeare|sonnets": [
        "Sonnets",
    ],
    # ── Vol 13: Virgil — Eclogues, Georgics, Aeneid (3 distinct GBWW entries) ──
    # These are listed as separate work rows in the workbook, not a corpus.
    # ── Vol 43: Hegel — Philosophy of Right, Philosophy of History ──
    "g. w. f. hegel|the philosophy of right": [
        "The Philosophy of Right",
    ],
    "g. w. f. hegel|the philosophy of history": [
        "The Philosophy of History",
    ],
    # ── Vol 29: Cervantes — Don Quixote (single work, multipart in GBWW) ──
    # ── Vol 30: Bacon — Advancement of Learning, Novum Organum, New Atlantis ──
    # ── Vol 32: Milton — English Minor Poems, Paradise Lost, Samson Agonistes, Areopagitica ──
    "john milton|english minor poems": [
        "On the Morning of Christ's Nativity",
        "A Paraphrase on Psalm 114",
        "Psalm 136",
        "The Passion",
        "On Time",
        "Upon the Circumcision",
        "At a Solemn Musick",
        "An Epitaph on the Marchioness of Winchester",
        "Song on May Morning",
        "On Shakespeare",
        "On the University Carrier",
        "Another on the Same",
        "L'Allegro",
        "Il Penseroso",
        "Sonnet to the Nightingale",
        "Sonnet on His Having Arrived at the Age of Twenty-Three",
        "Sonnet to the Lord General Cromwell",
        "Sonnet to Sir Henry Vane the Younger",
        "Sonnet on the Late Massacre in Piedmont",
        "Sonnet on His Blindness",
        "Sonnet to Cyriack Skinner",
        "Sonnet to the Memory of His Second Wife",
        "Arcades",
        "Comus",
        "Lycidas",
    ],
    # ── Vol 33: Pascal — Provincial Letters, Pensées, Scientific Treatises ──
    # ── Vol 33: Molière — The School for Wives, Critique, Tartuffe, Don Juan, Miser, Gentleman, Invalid ──
    # ── Vol 34: Newton — Principia, Opticks ──
    # ── Vol 35: Locke — Letter on Toleration, Second Treatise, Essay ──
    "john locke|a letter concerning toleration": [
        "A Letter Concerning Toleration",
    ],
    "john locke|second treatise of government": [
        "Second Treatise of Government",
    ],
    "john locke|an essay concerning human understanding": [
        "An Essay Concerning Human Understanding",
    ],
    # ── Vol 38: Montesquieu — The Spirit of Laws ──
    # ── Vol 38: Rousseau — Discourse on Inequality, Political Economy, Social Contract ──
    "jean-jacques rousseau|discourse on the origin of inequality": [
        "A Discourse on the Origin of Inequality",
    ],
    "jean-jacques rousseau|discourse on political economy": [
        "A Discourse on Political Economy",
    ],
    "jean-jacques rousseau|the social contract": [
        "The Social Contract",
    ],
    # ── Vol 39: Smith — Wealth of Nations ──
    # ── Vol 42: Kant — Critique of Pure Reason, Groundwork, Practical Reason, Metaphysics of Morals, Judgment ──
    # ── Vol 43: Mill — On Liberty, Representative Government, Utilitarianism ──
    # ── Vol 50: Marx — Communist Manifesto, Capital ──
    # ── Vol 53: Various — 20th-century foundations ──
    # Most entries are single-work; constituents are just the title itself.
}

# Works whose GBWW title is a corpus / grouping and should carry a non‑empty
# constituents array.  Key: `"<author>|<title>"`.
CORPUS_WORKS: set[str] = {
    "aeschylus|plays",
    "sophocles|plays",
    "euripides|plays",
    "aristophanes|plays",
    "plato|dialogues",
    "aristotle|works",
    "hippocrates|works",
    "archimedes|works",
    "plotinus|the six enneads",
    "william shakespeare|plays",
    "william shakespeare|sonnets",
    "john milton|english minor poems",
    "john locke|a letter concerning toleration",
    "john locke|second treatise of government",
    "john locke|an essay concerning human understanding",
    "jean-jacques rousseau|discourse on the origin of inequality",
    "jean-jacques rousseau|discourse on political economy",
    "jean-jacques rousseau|the social contract",
    "g. w. f. hegel|the philosophy of right",
    "g. w. f. hegel|the philosophy of history",
}


# ── helpers ────────────────────────────────────────────────────────────

def slugify(text: str) -> str:
    """Stable, reproducible slug: lowercased ASCII, spaces → hyphens."""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_ = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_.lower()).strip("-")


def normalize_key(author: str, title: str) -> str:
    """Key used to look up GBWW constituents."""
    return "|".join(
        unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        .strip()
        .lower()
        for s in (author, title)
    )


# ── extraction ─────────────────────────────────────────────────────────

def extract_works() -> list[dict]:
    """Read Audited Matches sheet, emit 183 normalized work rows."""
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Audited Matches"]

    # Verify expected count from Coverage Summary sheet
    cs = wb["Coverage Summary"]
    expected = None
    for row in cs.iter_rows(min_row=2, max_row=cs.max_row, values_only=True):
        if row[0] and str(row[0]).strip() == "Normalized works":
            expected = int(row[1])
            break

    works: list[dict] = []
    for idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2
    ):
        title_cell = (row[0] or "").strip()
        author_cell = (row[1] or "").strip()

        if not title_cell or not author_cell:
            continue  # skip empty rows

        work_id = f"{slugify(author_cell)}-{slugify(title_cell)}"

        # Determine constituents
        constituents: list[str] = []
        constituents_source: str = ""
        norm_key = normalize_key(author_cell, title_cell)

        if norm_key in CORPUS_WORKS:
            constituents = GBWW_CONSTITUENTS.get(norm_key, [])
            constituents_source = "GBWW second-edition contents (hardcoded)"

        entry: dict = {
            "workID": work_id,
            "row": idx,
            "author": author_cell,
            "title": title_cell,
            "constituents": constituents,
            "constituentsSource": constituents_source,
        }
        works.append(entry)

    if expected is not None and len(works) != expected:
        raise SystemExit(
            f"FATAL: expected {expected} normalized works, extracted {len(works)}"
        )

    print(f"Extracted {len(works)} works (expected {expected})")
    return works


def extract_seed() -> list[dict]:
    """Read Recording Versions sheet, emit verified seed rows."""
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Recording Versions"]

    seeds: list[dict] = []
    for idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2
    ):
        work_title = (row[0] or "").strip()
        author = (row[1] or "").strip()
        underlying_row = row[2]
        recording = (row[3] or "").strip()
        identifier = (row[8] or "").strip()
        match_class = (row[10] or "").strip()
        librivox_url = (row[7] or "").strip() if len(row) > 7 else ""

        if not work_title or not author:
            continue

        work_id = f"{slugify(author)}-{slugify(work_title)}"

        seeds.append(
            {
                "workID": work_id,
                "author": author,
                "title": work_title,
                "underlyingRow": underlying_row,
                "recordingTitle": recording,
                "identifier": identifier if identifier and identifier != "None" else "",
                "librivoxURL": librivox_url,
                "matchClass": match_class,
            }
        )

    print(f"Extracted {len(seeds)} verified seed rows")
    return seeds


# ── main ───────────────────────────────────────────────────────────────

def main() -> None:
    works = extract_works()
    seeds = extract_seed()

    # Write outputs
    json_opts = {"ensure_ascii": False, "indent": 2}

    OUT_WORKS.parent.mkdir(parents=True, exist_ok=True)
    OUT_WORKS.write_text(json.dumps(works, **json_opts) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_WORKS}")

    OUT_SEED.write_text(json.dumps(seeds, **json_opts) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_SEED}")


if __name__ == "__main__":
    main()
